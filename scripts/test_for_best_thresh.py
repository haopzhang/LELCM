import cv2
import numpy as np
import os
from scripts.mask_IoU import IoU
import torch.nn.functional as F
import torch
from utils import *
from skimage import measure
import pdb
from log import Log
from openpyxl import load_workbook

def target_edge(target_region, threshold):
    top_edge = list(target_region[0, :])  # 上边缘
    bottom_edge = list(target_region[-1, :])  # 下边缘
    left_edge = list(target_region[:, 0])  # 左边缘
    right_edge = list(target_region[:, -1])
    merged_edge = top_edge + bottom_edge + left_edge + right_edge
    for pixel in merged_edge:
        if pixel > threshold:
            return False
    return True


if __name__ == "__main__":
    imgs_path = "D:/111Project/data/SIRST/images/"
    masks_path = "D:/111Project/data/SIRST/masks/"
    mask_point_path = "D:/111Project/data/SIRST/masks_centroid/"
    img_names = os.listdir(imgs_path)
    log_fore = Log("test_4_best_thresh")

    wb = load_workbook('demo.xlsx')
    sheet = wb.active
    img_index = 1
    for index in range(0, 1):
        iou_sum = 0
        for img_name in img_names:
            img_index += 1
            img = cv2.imread(imgs_path + img_name, 0)
            m, n = img.shape
            ans_img = np.zeros((m, n))
            mask = cv2.imread(masks_path + img_name, 0)
            point_mask = cv2.imread(mask_point_path + img_name, 0)
            if img.shape != mask.shape or mask.shape != point_mask.shape:
                img_names.remove(img_name)
                continue


            label_point_mask = measure.label(point_mask)
            save_path = "./temp/"
            os.makedirs(save_path + img_name, exist_ok=True)
            target_index = 1
            for region in measure.regionprops(label_point_mask, cache=False):
                center = img[int(region.centroid[0]), int(region.centroid[1])]
                thresh = center*0.75
                core = 3
                # 扩充核大小，3 5 7 9...
                while True:
                    temp_img = np.zeros((m, n))
                    temp_img[int(region.centroid[0]), int(region.centroid[1])] = 255
                    nbr_mask = ((F.conv2d(torch.Tensor(temp_img).unsqueeze(0).unsqueeze(0),
                                          weight=torch.ones(1, 1, core, core), stride=1,
                                          padding=core // 2)) > 0).float()
                    point_region = measure.regionprops(np.squeeze(nbr_mask.cpu().numpy()).astype(np.uint8), cache=False)[0]
                    x1, y1, x2, y2 = point_region.bbox
                    if target_edge(img[x1:x2, y1:y2], 0.6*thresh) or core > 15:
                        cropped_img = np.squeeze(nbr_mask.cpu().numpy())*img
                        best_iou = 0
                        best_thresh = -1
                        for i in range(0, 1):
                            thresh = 0.8

                            candidate_pix = (cropped_img > thresh*255).astype(np.uint8)
                            label_candidate = measure.label(candidate_pix)
                            binary_image = (label_candidate == label_candidate[
                                int(region.centroid[0]), int(region.centroid[1])]).astype(np.uint8)
                            iou = IoU(binary_image*255, mask*np.squeeze(nbr_mask.cpu().numpy()))
                            if iou > best_iou:
                                best_iou = iou
                                best_thresh = thresh
                        if best_iou<0.5:
                            log_fore.log_print(img_name + ':' + str(best_iou) + '\t' + str(best_thresh) + '\t' + str(
                            np.std(img[x1:x2, y1:y2])) + '\t' + str(np.mean(img[x1:x2, y1:y2])) + '\t' + str(
                            center))
                            sheet.cell(row=img_index, column=1).value = img_name
                            sheet.cell(row=img_index, column=2).value = str(best_iou)
                            sheet.cell(row=img_index, column=3).value = str(best_thresh)
                            sheet.cell(row=img_index, column=4).value = np.std(img[x1:x2, y1:y2])
                            sheet.cell(row=img_index, column=5).value = np.mean(img[x1:x2, y1:y2])
                            sheet.cell(row=img_index, column=6).value = center
                            wb.save('demo.xlsx')
                        ans_img += binary_image
                        break

                    else:
                        core += 2
                cv2.imwrite(save_path + img_name + "/region"+str(target_index)+".png", img[x1:x2, y1:y2])
                target_index += 1
                cv2.imwrite(save_path + img_name + "/region_out" + str(target_index) + ".png", binary_image*255)

            cv2.imwrite(save_path + img_name + "/image.png", img)
            cv2.imwrite(save_path + img_name + "/ans.png", ans_img*255)
            cv2.imwrite(save_path + img_name + "/mask.png", mask)

            iou = IoU(ans_img*255, mask)
            print(iou)
            if iou < 0.5:
                log_fore.log_print(img_name)


            iou_sum += iou
            # print(iou)
        print("mIoU:"+str(index)+":"+str(iou_sum/len(img_names)))




