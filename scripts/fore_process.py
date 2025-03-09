import cv2
import numpy as np
import os
import torch.nn.functional as F
import torch
from utils import *
from skimage import measure
import pdb
from log import Log
from openpyxl import load_workbook
from mask_IoU import *


def target_edge(target_region, threshold):
    top_edge = list(target_region[0, :])  # 上边缘
    bottom_edge = list(target_region[-1, :])  # 下边缘
    left_edge = list(target_region[:, 0])  # 左边缘
    right_edge = list(target_region[:, -1])
    merged_edge = top_edge + bottom_edge + left_edge + right_edge
    for pixel in merged_edge:
        if pixel > threshold:
            return False
    return True


if __name__ == "__main__":
    imgs_path = "D:/111Project/data/SIRST/images/"
    masks_path = "D:/111Project/data/SIRST/masks/"
    mask_point_path = "D:/111Project/data/ans/SIRST/TLLCM_Try/"
    save_path = "D:/111Project/data/SIRST/masks_TLLCM_try/"
    os.makedirs(save_path, exist_ok=True)
    img_names = os.listdir(imgs_path)
    log_fore = Log("bad_images_list")
    wb = load_workbook('demo.xlsx')
    sheet = wb.active
    img_index = 0

    for index in range(0, 1):
        iou_sum = 0
        pd_sum = 0
        f_pixel_sum = 0
        all_pix = 0
        for img_name in img_names:
            img_index += 1
            img = cv2.imread(imgs_path + img_name, 0)
            m, n = img.shape
            ans_img = np.zeros((m, n))
            mask = cv2.imread(masks_path + img_name, 0)
            point_mask = cv2.imread(mask_point_path + img_name, 0)
            if img.shape != mask.shape or mask.shape != point_mask.shape:
                img_names.remove(img_name)
                cv2.imwrite(save_path + img_name, mask)
                continue

            label_point_mask = measure.label(point_mask)
            # save_path = "./temp/"
            # os.makedirs(save_path + img_name, exist_ok=True)
            target_index = 0
            for region in measure.regionprops(label_point_mask, cache=False):
                center = img[int(region.centroid[0]), int(region.centroid[1])]
                core = 3
                # 扩充核大小，3 5 7 9...
                while True:
                    temp_img = np.zeros((m, n))
                    temp_img[int(region.centroid[0]), int(region.centroid[1])] = 255
                    nbr_mask = ((F.conv2d(torch.Tensor(temp_img).unsqueeze(0).unsqueeze(0),
                                          weight=torch.ones(1, 1, core, core), stride=1,
                                          padding=core // 2)) > 0).float()
                    point_region = measure.regionprops(np.squeeze(nbr_mask.cpu().numpy()).astype(np.uint8), cache=False)[0]
                    x1, y1, x2, y2 = point_region.bbox
                    if target_edge(img[x1:x2, y1:y2], 0.8*center) or core > 15:
                        thresh = center - np.std(img[x1:x2, y1:y2])
                        # print(thresh)
                        candidate_pix = (np.squeeze(nbr_mask.cpu().numpy())*img > thresh).astype(np.uint8)
                        label_candidate = measure.label(candidate_pix)
                        binary_image = (label_candidate == label_candidate[int(region.centroid[0]), int(region.centroid[1])]).astype(np.uint8)

                        ans_img += binary_image
                        break

                    else:
                        core += 2
            '''
                cv2.imwrite(save_path + img_name + "/region"+str(target_index)+".png", img[x1:x2, y1:y2])
                target_index += 1
                cv2.imwrite(save_path + img_name + "/region_out" + str(target_index) + ".png", binary_image*255)

            cv2.imwrite(save_path + img_name + "/image.png", img)
            cv2.imwrite(save_path + img_name + "/ans.png", ans_img*255)
            cv2.imwrite(save_path + img_name + "/mask.png", mask)
            '''
            iou = IoU((ans_img*255).astype(np.uint8), mask)
            pd = Pd((ans_img*255).astype(np.uint8), mask)
            fa, pix_mask = Fa((ans_img*255).astype(np.uint8), mask)
            # save_path = "D:/111Project/data/SIRST/masks_preprocess/"
            cv2.imwrite(save_path + img_name, ans_img*255)
            '''
            if iou < 0.5:
                log_fore.log_print(img_name)
            if iou < 0.5:
                sheet.cell(row=img_index, column=1).value = img_name
                sheet.cell(row=img_index, column=2).value = str(iou)
                sheet.cell(row=img_index, column=4).value = np.std(img[x1:x2, y1:y2])
                sheet.cell(row=img_index, column=5).value = np.mean(img[x1:x2, y1:y2])
                sheet.cell(row=img_index, column=6).value = center
                wb.save('demo.xlsx')
            '''

            iou_sum += iou
            pd_sum += pd
            f_pixel_sum += fa
            all_pix += pix_mask
            # print(iou)
        print("mIoU:"+str(index)+":"+str(iou_sum/len(img_names)))
        print("Pd:" + str(index) + ":" + str(pd_sum / len(img_names)))
        print("Fa:" + str(index) + ":" + str(f_pixel_sum/all_pix))



